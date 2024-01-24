import openpyxl as xl
from openpyxl.chart import BarChart, Reference

wb = xl.load_workbook('transactions.xlsx') # load the workbook
sheet = wb['Sheet1'] # access the sheet
cell = sheet['A1']
cell = sheet.cell(1,1) #row, col

for row in range(2, sheet.max_row+1):
    cell = sheet.cell(row, 3)
    correctPrice = cell.value * 0.9 #reduce 10%
    correctPriceCell = sheet.cell(row, 4)
    correctPriceCell.value = correctPrice

values = Reference(sheet, min_row = 2, max_row = sheet.max_row, min_col = 4, max_col = 4)

chart = BarChart()
chart.add_data(values)
sheet.add_chart(chart, 'b6')
wb.save('transactions2.xlsx')