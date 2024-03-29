import openpyxl as xl
from openpyxl.chart import BarChart, Reference

wb = xl.load_workbook('transactions.xlsx') # load the workbook
sheet = wb['Sheet1'] # access the sheet
cell = sheet['A1']
cell = sheet.cell(1,1) #row, col
values = Reference(sheet, min_row = 2, max_row = sheet.max_row, min_col = 4, max_col = 4)
print(values)